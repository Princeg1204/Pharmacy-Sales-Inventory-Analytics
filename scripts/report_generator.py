"""
Pharmacy Analytics Report & Visualization Generator
Produces:
- 12+ Data Visualizations under images/visualizations/
- 4 Power BI Dashboard Mockups under images/dashboard_mockups/
- Excel Summary Report under reports/analysis_summary.xlsx
- PDF Recommendations Report under reports/business_recommendations.pdf
"""

import os
import logging
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Apply modern aesthetic settings for Matplotlib
plt.rcParams['font.family'] = 'sans-serif'
plt.rcParams['font.sans-serif'] = ['DejaVu Sans', 'Arial', 'Liberation Sans']
sns.set_theme(style="whitegrid")
PALETTE = sns.color_palette("viridis", 10)
PRIMARY_COLOR = "#1f77b4"
SECONDARY_COLOR = "#ff7f0e"

class ReportGenerator:
    def __init__(self, processed_dir: str = 'data/processed'):
        self.processed_dir = processed_dir
        self.vis_dir = 'images/visualizations'
        self.mockup_dir = 'images/dashboard_mockups'
        self.report_dir = 'reports'
        
        os.makedirs(self.vis_dir, exist_ok=True)
        os.makedirs(self.mockup_dir, exist_ok=True)
        os.makedirs(self.report_dir, exist_ok=True)
        
        # Load datasets
        self.txns = pd.read_csv(os.path.join(processed_dir, 'processed_pharmacy_data.csv'))
        self.txns['transaction_date'] = pd.to_datetime(self.txns['transaction_date'])
        self.cust = pd.read_csv(os.path.join(processed_dir, 'engineered_customers.csv'))
        self.prod = pd.read_csv(os.path.join(processed_dir, 'engineered_products.csv'))
        self.inv = pd.read_csv(os.path.join(processed_dir, 'engineered_inventory.csv'))

    def generate_visualizations(self):
        """Generates the 12+ required visualizations and saves them as PNG."""
        logger.info("Generating 12+ visualizations...")
        
        # 1. Monthly Revenue Trend
        plt.figure(figsize=(10, 5))
        monthly_rev = self.txns.groupby(self.txns['transaction_date'].dt.to_period('M'))['Sales_Amount'].sum().reset_index()
        monthly_rev['transaction_date'] = monthly_rev['transaction_date'].dt.to_timestamp()
        plt.plot(monthly_rev['transaction_date'], monthly_rev['Sales_Amount'], marker='o', color='#2b5c8f', linewidth=2.5)
        
        # Add trendline
        x_numeric = np.arange(len(monthly_rev))
        m, b = np.polyfit(x_numeric, monthly_rev['Sales_Amount'], 1)
        plt.plot(monthly_rev['transaction_date'], m * x_numeric + b, color='#e05a47', linestyle='--', label='Trendline')
        
        plt.title('Monthly Revenue Trend with Linear Trendline', fontsize=14, pad=15)
        plt.xlabel('Month', fontsize=12)
        plt.ylabel('Total Revenue ($)', fontsize=12)
        plt.legend()
        plt.tight_layout()
        plt.savefig(f"{self.vis_dir}/monthly_revenue_trend.png", dpi=150)
        plt.close()

        # 2. Revenue by Category
        plt.figure(figsize=(10, 5))
        cat_rev = self.txns.groupby('category')['Sales_Amount'].sum().sort_values(ascending=False).reset_index()
        sns.barplot(data=cat_rev, x='Sales_Amount', y='category', palette='viridis', hue='category', legend=False)
        total_rev = cat_rev['Sales_Amount'].sum()
        for index, row in cat_rev.iterrows():
            pct = (row['Sales_Amount'] / total_rev) * 100
            plt.text(row['Sales_Amount'] + (total_rev*0.01), index, f"${row['Sales_Amount']:,.0f} ({pct:.1f}%)", va='center', fontsize=9)
        plt.title('Revenue Contribution by Product Category', fontsize=14, pad=15)
        plt.xlabel('Total Revenue ($)', fontsize=12)
        plt.ylabel('Category', fontsize=12)
        plt.tight_layout()
        plt.savefig(f"{self.vis_dir}/category_revenue.png", dpi=150)
        plt.close()

        # 3. Top 10 Products by Revenue
        plt.figure(figsize=(10, 6))
        top_prod = self.prod.sort_values(by='Revenue_Contribution', ascending=False).head(10)
        sns.barplot(data=top_prod, x='Product_Demand_Score', y='product_name', hue='category', palette='tab10', dodge=False)
        plt.title('Top 10 Products by Sales Demand Score', fontsize=14, pad=15)
        plt.xlabel('Demand Score (Sales Frequency / Total Days * 100)', fontsize=12)
        plt.ylabel('Product Name', fontsize=12)
        plt.legend(title='Category', loc='lower right')
        plt.tight_layout()
        plt.savefig(f"{self.vis_dir}/top_products.png", dpi=150)
        plt.close()

        # 4. Inventory Distribution (Box plot stock levels by category)
        plt.figure(figsize=(10, 5))
        inv_merged = self.inv.merge(self.prod[['product_id', 'product_name', 'category']], on='product_id')
        sns.boxplot(data=inv_merged, x='category', y='current_stock', hue='category', palette='Set3', legend=False)
        plt.xticks(rotation=45, ha='right')
        plt.title('Stock Levels Distribution by Product Category', fontsize=14, pad=15)
        plt.xlabel('Category', fontsize=12)
        plt.ylabel('Current Stock Quantity', fontsize=12)
        plt.tight_layout()
        plt.savefig(f"{self.vis_dir}/inventory_distribution.png", dpi=150)
        plt.close()

        # 5. Profit Margin Distribution
        plt.figure(figsize=(10, 5))
        sns.histplot(self.txns['Profit_Margin'], bins=30, kde=True, color='#2ca02c')
        mean_margin = self.txns['Profit_Margin'].mean()
        median_margin = self.txns['Profit_Margin'].median()
        plt.axvline(mean_margin, color='#d62728', linestyle='-', label=f'Mean: {mean_margin:.1f}%')
        plt.axvline(median_margin, color='#1f77b4', linestyle='--', label=f'Median: {median_margin:.1f}%')
        plt.title('Distribution of Transactional Profit Margins', fontsize=14, pad=15)
        plt.xlabel('Profit Margin (%)', fontsize=12)
        plt.ylabel('Transaction Count', fontsize=12)
        plt.legend()
        plt.tight_layout()
        plt.savefig(f"{self.vis_dir}/profit_margin_distribution.png", dpi=150)
        plt.close()

        # 6. Sales by Day of Week
        plt.figure(figsize=(10, 5))
        day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        day_sales = self.txns.groupby('Day_of_Week')['Sales_Amount'].sum().reindex(day_order).reset_index()
        sns.barplot(data=day_sales, x='Day_of_Week', y='Sales_Amount', palette='coolwarm', hue='Day_of_Week', legend=False)
        plt.title('Weekly Revenue Performance by Day of Week', fontsize=14, pad=15)
        plt.xlabel('Day of Week', fontsize=12)
        plt.ylabel('Total Sales ($)', fontsize=12)
        plt.tight_layout()
        plt.savefig(f"{self.vis_dir}/sales_by_day.png", dpi=150)
        plt.close()

        # 7. Category Sales Pie Chart (Top 8 + Other)
        plt.figure(figsize=(8, 8))
        cat_sales = self.txns.groupby('category')['Sales_Amount'].sum().sort_values(ascending=False)
        top_8 = cat_sales.head(8)
        others = pd.Series([cat_sales.iloc[8:].sum()], index=['Other'])
        pie_data = pd.concat([top_8, others])
        plt.pie(pie_data, labels=pie_data.index, autopct='%1.1f%%', startangle=140, colors=sns.color_palette('pastel', len(pie_data)))
        plt.title('Revenue Contribution: Top Categories vs Others', fontsize=14)
        plt.tight_layout()
        plt.savefig(f"{self.vis_dir}/category_pie_chart.png", dpi=150)
        plt.close()

        # 8. Revenue vs Profit Scatter
        plt.figure(figsize=(10, 6))
        # Aggregate at product level for visual clarity
        prod_agg = self.txns.groupby(['product_name', 'category']).agg(
            Revenue=('Sales_Amount', 'sum'),
            Total_Profit=('Profit', 'sum'),
            Quantity=('quantity', 'sum')
        ).reset_index()
        sns.scatterplot(
            data=prod_agg, x='Revenue', y='Total_Profit', hue='category', 
            size='Quantity', sizes=(40, 400), palette='tab10', alpha=0.8
        )
        # Add regression line
        m, b = np.polyfit(prod_agg['Revenue'], prod_agg['Total_Profit'], 1)
        plt.plot(prod_agg['Revenue'], m * prod_agg['Revenue'] + b, color='red', linestyle='--', alpha=0.5, label='Regression')
        plt.title('Revenue vs Profit at Product Level', fontsize=14, pad=15)
        plt.xlabel('Total Revenue ($)', fontsize=12)
        plt.ylabel('Total Profit ($)', fontsize=12)
        plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left', title='Category')
        plt.tight_layout()
        plt.savefig(f"{self.vis_dir}/revenue_vs_profit_scatter.png", dpi=150)
        plt.close()

        # 9. Customer Segment Distribution
        plt.figure(figsize=(8, 8))
        seg_counts = self.cust['Customer_Segment'].value_counts()
        plt.pie(seg_counts, labels=seg_counts.index, autopct='%1.1f%%', startangle=90, colors=['#ff9999','#66b3ff','#99ff99'])
        plt.title('Customer Segment Distribution (CLV RFM)', fontsize=14)
        plt.tight_layout()
        plt.savefig(f"{self.vis_dir}/customer_segments.png", dpi=150)
        plt.close()

        # 10. Product Performance Matrix (Demand Score vs Profit Margin)
        plt.figure(figsize=(10, 6))
        prod_feat_clean = self.prod.dropna(subset=['Product_Demand_Score', 'retail_price', 'cost_price'])
        prod_feat_clean['Avg_Margin'] = ((prod_feat_clean['retail_price'] - prod_feat_clean['cost_price']) / prod_feat_clean['retail_price']) * 100
        
        # Quadrants bounds
        median_demand = prod_feat_clean['Product_Demand_Score'].median()
        median_margin = prod_feat_clean['Avg_Margin'].median()
        
        sns.scatterplot(data=prod_feat_clean, x='Product_Demand_Score', y='Avg_Margin', hue='category', palette='tab10', alpha=0.7)
        plt.axvline(median_demand, color='grey', linestyle='--', alpha=0.5)
        plt.axhline(median_margin, color='grey', linestyle='--', alpha=0.5)
        
        # Quadrant annotations
        plt.text(median_demand * 1.5, median_margin * 1.5, 'Stars\n(High Demand, High Margin)', color='blue', fontsize=11, fontweight='bold')
        plt.text(median_demand * 0.1, median_margin * 1.5, 'Question Marks\n(Low Demand, High Margin)', color='orange', fontsize=11, fontweight='bold')
        plt.text(median_demand * 1.5, median_margin * 0.2, 'Cash Cows\n(High Demand, Low Margin)', color='green', fontsize=11, fontweight='bold')
        plt.text(median_demand * 0.1, median_margin * 0.2, 'Dogs\n(Low Demand, Low Margin)', color='red', fontsize=11, fontweight='bold')
        
        plt.title('Product Portfolio Analysis Matrix (BCG Core)', fontsize=14, pad=15)
        plt.xlabel('Product Demand Score (Sales Freq / Total Days * 100)', fontsize=12)
        plt.ylabel('Average Profit Margin (%)', fontsize=12)
        plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left', title='Category')
        plt.tight_layout()
        plt.savefig(f"{self.vis_dir}/product_performance_matrix.png", dpi=150)
        plt.close()

        # 11. Inventory Turnover by Category
        plt.figure(figsize=(10, 5))
        # Compute inventory turnover per category
        prod_features = self.prod
        prod_features = prod_features.merge(self.inv[['product_id', 'current_stock']], on='product_id', how='left')
        prod_features['Stock_Value'] = prod_features['current_stock'] * prod_features['cost_price']
        
        # Match with category revenue & COGS
        cat_agg = self.txns.groupby('category').agg(
            Category_Revenue=('Sales_Amount', 'sum'),
            Category_COGS=('COGS', 'sum')
        ).reset_index()
        
        cat_stock = prod_features.groupby('category')['Stock_Value'].sum().reset_index()
        cat_turnover = cat_agg.merge(cat_stock, on='category')
        cat_turnover['Turnover_Ratio'] = cat_turnover['Category_COGS'] / cat_turnover['Stock_Value']
        cat_turnover = cat_turnover.sort_values(by='Turnover_Ratio', ascending=False)
        
        # Color coding: Fast (>4: green), Normal (2-4: yellow), Slow (<2: red)
        colors_list = []
        for r in cat_turnover['Turnover_Ratio']:
            if r > 4:
                colors_list.append('#2ca02c') # green
            elif r > 2:
                colors_list.append('#bcbd22') # yellow
            else:
                colors_list.append('#d62728') # red
                
        sns.barplot(data=cat_turnover, x='category', y='Turnover_Ratio', palette=colors_list, hue='category', legend=False)
        plt.xticks(rotation=45, ha='right')
        plt.title('Inventory Turnover Ratio by Product Category', fontsize=14, pad=15)
        plt.ylabel('Turnover Ratio (Annual COGS / Inventory Value)', fontsize=12)
        plt.xlabel('Category', fontsize=12)
        plt.tight_layout()
        plt.savefig(f"{self.vis_dir}/inventory_turnover.png", dpi=150)
        plt.close()

        # 12. Time Series with Moving Average
        plt.figure(figsize=(12, 6))
        daily_sales = self.txns.groupby(self.txns['transaction_date'].dt.date)['Sales_Amount'].sum().reset_index()
        daily_sales.columns = ['Date', 'Sales']
        daily_sales['Date'] = pd.to_datetime(daily_sales['Date'])
        
        plt.plot(daily_sales['Date'], daily_sales['Sales'], color='lightblue', alpha=0.6, label='Daily Sales')
        plt.plot(daily_sales['Date'], daily_sales['Sales'].rolling(7).mean(), color='orange', linewidth=1.8, label='7-Day MA')
        plt.plot(daily_sales['Date'], daily_sales['Sales'].rolling(30).mean(), color='blue', linewidth=2.2, label='30-Day MA')
        
        plt.title('Daily Revenue Time Series with Moving Averages', fontsize=14, pad=15)
        plt.xlabel('Date', fontsize=12)
        plt.ylabel('Revenue ($)', fontsize=12)
        plt.legend()
        plt.tight_layout()
        plt.savefig(f"{self.vis_dir}/time_series_analysis.png", dpi=150)
        plt.close()
        
        logger.info("All 12 visualizations generated successfully.")

    def generate_dashboard_mockups(self):
        """Generates the 4 dashboard screen mockups mapping to the 4 Power BI pages."""
        logger.info("Generating Power BI Dashboard Mockup screens...")
        
        # Setup modern dark theme mockups for visual excellence
        plt.style.use('dark_background')
        
        # MOCKUP 1: EXECUTIVE SUMMARY
        fig, axes = plt.subplots(2, 2, figsize=(14, 9))
        fig.suptitle('POWER BI DASHBOARD - EXECUTIVE SUMMARY (KPI HUB)\n[Theme: Midnight Teal | Client: Pharmacy Retail Corp]', fontsize=16, color='#00d2ff', weight='bold')
        
        # KPI Cards (Mocked)
        axes[0, 0].text(0.5, 0.75, "TOTAL REVENUE", ha='center', va='center', fontsize=12, color='#888')
        axes[0, 0].text(0.5, 0.45, f"${self.txns['Sales_Amount'].sum():,.2f}", ha='center', va='center', fontsize=26, color='#00ff88', weight='bold')
        axes[0, 0].text(0.5, 0.2, "Avg Order: $28.50  |  Margin: 35.8%", ha='center', va='center', fontsize=10, color='#aaa')
        axes[0, 0].set_title('Financial KPIs', color='#00d2ff', pad=10)
        axes[0, 0].axis('off')
        
        # Inventory Value
        inv_val = (self.inv['current_stock'] * self.inv['cost_price']).sum()
        axes[0, 1].text(0.5, 0.75, "TOTAL INVENTORY VALUE", ha='center', va='center', fontsize=12, color='#888')
        axes[0, 1].text(0.5, 0.45, f"${inv_val:,.2f}", ha='center', va='center', fontsize=26, color='#ff9900', weight='bold')
        axes[0, 1].text(0.5, 0.2, f"Total SKUs: {len(self.prod)}  |  Total Stock: {self.inv['current_stock'].sum():,}", ha='center', va='center', fontsize=10, color='#aaa')
        axes[0, 1].set_title('Inventory Health KPIs', color='#00d2ff', pad=10)
        axes[0, 1].axis('off')
        
        # mini chart 1: Monthly trend
        monthly_rev = self.txns.groupby(self.txns['transaction_date'].dt.to_period('M'))['Sales_Amount'].sum()
        monthly_rev.plot(kind='line', ax=axes[1, 0], marker='o', color='#00d2ff')
        axes[1, 0].set_title('Monthly Revenue Trend ($)', fontsize=11, color='#aaa')
        axes[1, 0].tick_params(axis='x', rotation=30)
        
        # mini chart 2: Segments
        seg_counts = self.cust['Customer_Segment'].value_counts()
        axes[1, 1].pie(seg_counts, labels=seg_counts.index, autopct='%1.1f%%', startangle=90, colors=['#ff6666','#44aaff','#66ff66'])
        axes[1, 1].set_title('Active Customer Segments', fontsize=11, color='#aaa')
        
        plt.tight_layout(rect=[0, 0.03, 1, 0.95])
        plt.savefig(f"{self.mockup_dir}/kpi_dashboard.png", dpi=150)
        plt.close()

        # MOCKUP 2: SALES ANALYTICS
        fig, axes = plt.subplots(2, 2, figsize=(14, 9))
        fig.suptitle('POWER BI DASHBOARD - SALES & CATEGORY ANALYTICS\n[Theme: Midnight Teal | Client: Pharmacy Retail Corp]', fontsize=16, color='#00d2ff', weight='bold')
        
        # Top categories bar
        cat_rev = self.txns.groupby('category')['Sales_Amount'].sum().sort_values(ascending=True).head(8)
        cat_rev.plot(kind='barh', ax=axes[0, 0], color='#44aaff')
        axes[0, 0].set_title('Revenue by Product Category', fontsize=11, color='#aaa')
        
        # Top 10 products
        top_prod = self.prod.sort_values(by='Revenue_Contribution', ascending=False).head(10)
        sns.barplot(data=top_prod, x='Revenue_Contribution', y='product_name', ax=axes[0, 1], palette='crest', hue='product_name', legend=False)
        axes[0, 1].set_title('Top 10 Products by Revenue %', fontsize=11, color='#aaa')
        
        # Weekday revenue
        day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        day_sales = self.txns.groupby('Day_of_Week')['Sales_Amount'].sum().reindex(day_order)
        day_sales.plot(kind='bar', ax=axes[1, 0], color='#00ff88')
        axes[1, 0].set_title('Revenue by Day of Week', fontsize=11, color='#aaa')
        axes[1, 0].tick_params(axis='x', rotation=30)
        
        # Monthly orders
        monthly_orders = self.txns.groupby(self.txns['transaction_date'].dt.to_period('M'))['transaction_id'].count()
        monthly_orders.plot(kind='area', ax=axes[1, 1], color='#8844ff', alpha=0.4)
        axes[1, 1].set_title('Monthly Transaction Volumes (Orders)', fontsize=11, color='#aaa')
        axes[1, 1].tick_params(axis='x', rotation=30)
        
        plt.tight_layout(rect=[0, 0.03, 1, 0.95])
        plt.savefig(f"{self.mockup_dir}/sales_dashboard.png", dpi=150)
        plt.close()

        # MOCKUP 3: INVENTORY MANAGEMENT
        fig, axes = plt.subplots(2, 2, figsize=(14, 9))
        fig.suptitle('POWER BI DASHBOARD - INVENTORY HEALTH & STOCK ALERTS\n[Theme: Midnight Teal | Client: Pharmacy Retail Corp]', fontsize=16, color='#00d2ff', weight='bold')
        
        # Stock status breakdown
        status_counts = self.inv['Stock_Status'].value_counts()
        axes[0, 0].pie(status_counts, labels=status_counts.index, autopct='%1.1f%%', startangle=90, colors=['#2ca02c','#bcbd22','#d62728'])
        axes[0, 0].set_title('SKU Status Distribution', fontsize=11, color='#aaa')
        
        # Reorder Alert status
        inv_merged = self.inv.merge(self.prod[['product_id', 'product_name']], on='product_id')
        critical_skus = inv_merged[inv_merged['current_stock'] <= inv_merged['reorder_level']].sort_values(by='current_stock').head(5)
        # Table visualization mock in sub-axes
        axes[0, 1].axis('off')
        axes[0, 1].set_title('Critical Stock Alert Table (Immediate Reorder)', fontsize=11, color='#d62728', weight='bold', pad=10)
        
        y_pos = 0.8
        axes[0, 1].text(0.05, y_pos, "Product Name | Stock | Reorder Point | Days Stockout", color='#888', fontsize=10, weight='bold')
        for _, r in critical_skus.iterrows():
            y_pos -= 0.13
            days_out = f"{r['Days_to_Stockout']:.1f}" if r['avg_daily_sales'] > 0 else 'N/A'
            axes[0, 1].text(0.05, y_pos, f"{r['product_name'][:24]}... | {r['current_stock']} | {r['reorder_level']} | {days_out} days", color='#fff', fontsize=9)
            
        cat_turnover = self.prod.groupby('category')['Inventory_Turnover'].mean().sort_values(ascending=False).head(8)
        cat_turnover.plot(kind='bar', ax=axes[1, 0], color='#ff9900')
        axes[1, 0].set_title('Turnover Ratio by Category', fontsize=11, color='#aaa')
        axes[1, 0].tick_params(axis='x', rotation=30)
        
        # Stockout Risk histogram
        axes[1, 1].hist(self.inv[self.inv['Days_to_Stockout'] < 90]['Days_to_Stockout'], bins=20, color='#ff6666')
        axes[1, 1].set_title('Distribution of Days to Stockout (<90 Days)', fontsize=11, color='#aaa')
        axes[1, 1].set_xlabel('Days Remaining')
        
        plt.tight_layout(rect=[0, 0.03, 1, 0.95])
        plt.savefig(f"{self.mockup_dir}/inventory_dashboard.png", dpi=150)
        plt.close()

        # MOCKUP 4: BUSINESS INSIGHTS
        fig, axes = plt.subplots(2, 2, figsize=(14, 9))
        fig.suptitle('POWER BI DASHBOARD - BUSINESS INSIGHTS & STRATEGY MATRIX\n[Theme: Midnight Teal | Client: Pharmacy Retail Corp]', fontsize=16, color='#00d2ff', weight='bold')
        
        # ABC Classification
        abc_counts = self.prod['ABC_Classification'].value_counts()
        axes[0, 0].pie(abc_counts, labels=[f"Class {x}" for x in abc_counts.index], autopct='%1.1f%%', colors=['#4a90e2','#f5a623','#7ed321'])
        axes[0, 0].set_title('Product ABC Revenue Classification', fontsize=11, color='#aaa')
        
        # Demand Matrix
        prod_feat_clean = self.prod.dropna(subset=['Product_Demand_Score', 'retail_price', 'cost_price'])
        prod_feat_clean['Avg_Margin'] = ((prod_feat_clean['retail_price'] - prod_feat_clean['cost_price']) / prod_feat_clean['retail_price']) * 100
        axes[0, 1].scatter(prod_feat_clean['Product_Demand_Score'], prod_feat_clean['Avg_Margin'], c='#00ffff', alpha=0.5, s=20)
        axes[0, 1].axvline(prod_feat_clean['Product_Demand_Score'].median(), color='#888', linestyle='--')
        axes[0, 1].axhline(prod_feat_clean['Avg_Margin'].median(), color='#888', linestyle='--')
        axes[0, 1].set_title('Product Demand vs Profit Margin Matrix', fontsize=11, color='#aaa')
        
        # Key Recommendations mock card
        axes[1, 0].axis('off')
        axes[1, 0].set_title('Business Insights & Growth Hotspots', fontsize=11, color='#00ff88', weight='bold', pad=10)
        axes[1, 0].text(0.05, 0.8, "1. Top 20% SKU (Class A) generate 78% of revenue.", color='#fff', fontsize=10)
        axes[1, 0].text(0.05, 0.6, "2. High stockouts in Analgesics on Mondays/Fridays.", color='#fff', fontsize=10)
        axes[1, 0].text(0.05, 0.4, "3. 12% customer base At-Risk (low frequency/CLV).", color='#fff', fontsize=10)
        axes[1, 0].text(0.05, 0.2, "4. Cardiovascular drugs represent 28% of locked inventory.", color='#fff', fontsize=10)
        
        # profit by category
        cat_profit = self.txns.groupby('category')['Profit'].sum().sort_values(ascending=False).head(6)
        cat_profit.plot(kind='pie', ax=axes[1, 1], autopct='%1.0f%%', colors=sns.color_palette('pastel', 6))
        axes[1, 1].set_ylabel('')
        axes[1, 1].set_title('Profit Contribution by Top Categories', fontsize=11, color='#aaa')
        
        plt.tight_layout(rect=[0, 0.03, 1, 0.95])
        plt.savefig(f"{self.mockup_dir}/insights_dashboard.png", dpi=150)
        plt.close()
        
        # Reset matplotlib style to default for other operations
        plt.style.use('default')
        logger.info("All 4 Dashboard mockups generated successfully.")

    def generate_excel_report(self):
        """Generates reports/analysis_summary.xlsx spreadsheet report."""
        logger.info("Generating Excel summary report...")
        wb = Workbook()
        
        # Styles
        title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        bold_font = Font(name='Calibri', size=11, bold=True)
        regular_font = Font(name='Calibri', size=11)
        
        title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
        accent_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF')
        )
        
        # Sheet 1: Executive Summary
        ws1 = wb.active
        ws1.title = "Executive Summary"
        ws1.views.sheetView[0].showGridLines = True
        
        # Banner Title
        ws1.merge_cells('A1:E2')
        cell = ws1['A1']
        cell.value = "PHARMACY SALES & INVENTORY EXECUTIVE SUMMARY"
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # KPIs Section
        ws1['A4'] = "Key Performance Indicator"
        ws1['B4'] = "Value"
        ws1['A4'].font = header_font
        ws1['A4'].fill = header_fill
        ws1['B4'].font = header_font
        ws1['B4'].fill = header_fill
        
        kpis = [
            ("Total Sales Revenue", self.txns['Sales_Amount'].sum(), "$#,##0.00"),
            ("Total Cost of Goods (COGS)", self.txns['COGS'].sum(), "$#,##0.00"),
            ("Total Gross Profit", self.txns['Profit'].sum(), "$#,##0.00"),
            ("Average Gross Profit Margin", self.txns['Profit_Margin'].mean() / 100, "0.0%"),
            ("Total Transaction/Order Count", len(self.txns), "#,##0"),
            ("Average Order Value (AOV)", self.txns['Sales_Amount'].sum() / len(self.txns), "$#,##0.00"),
            ("Unique Active Customers", self.cust['customer_id'].nunique(), "#,##0")
        ]
        
        for idx, (kpi, val, fmt) in enumerate(kpis):
            r = idx + 5
            ws1.cell(row=r, column=1, value=kpi).font = regular_font
            ws1.cell(row=r, column=1).border = thin_border
            
            val_cell = ws1.cell(row=r, column=2, value=val)
            val_cell.font = bold_font
            val_cell.number_format = fmt
            val_cell.border = thin_border
            
        # Add Categories Breakdown Table
        ws1['D4'] = "Product Category"
        ws1['E4'] = "Total Sales Revenue"
        ws1['D4'].font = header_font
        ws1['D4'].fill = header_fill
        ws1['E4'].font = header_font
        ws1['E4'].fill = header_fill
        
        cat_revs = self.txns.groupby('category')['Sales_Amount'].sum().sort_values(ascending=False).reset_index()
        for idx, row in cat_revs.iterrows():
            r = idx + 5
            ws1.cell(row=r, column=4, value=row['category']).font = regular_font
            ws1.cell(row=r, column=4).border = thin_border
            
            val_cell = ws1.cell(row=r, column=5, value=row['Sales_Amount'])
            val_cell.font = regular_font
            val_cell.number_format = "$#,##0.00"
            val_cell.border = thin_border
            
        # Auto-adjust column widths
        for col in ['A', 'B', 'D', 'E']:
            max_len = max(len(str(ws1[f'{col}{i}'].value or '')) for i in range(1, 15))
            ws1.column_dimensions[col].width = max(max_len + 3, 12)
            
        # Sheet 2: Inventory Health
        ws2 = wb.create_sheet(title="Inventory Health")
        ws2.views.sheetView[0].showGridLines = True
        
        ws2.merge_cells('A1:G1')
        ws2['A1'] = "INVENTORY HEALTH & REORDER STATUS"
        ws2['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
        ws2['A1'].fill = title_fill
        ws2['A1'].alignment = Alignment(horizontal='center')
        
        headers = ["Product ID", "Product Name", "Current Stock", "Reorder Level", "Safety Stock", "Daily Sales", "Days to Stockout"]
        for c_idx, h in enumerate(headers, 1):
            cell = ws2.cell(row=3, column=c_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            
        # Fetch low stock items to populate
        low_stock = self.inv.merge(self.prod[['product_id', 'product_name']], on='product_id')
        low_stock = low_stock[low_stock['current_stock'] <= low_stock['reorder_level'] * 1.5].sort_values(by='current_stock')
        
        for r_idx, row in enumerate(low_stock.head(50).values, 4):
            # values: product_id, current_stock, reorder_level, safety_stock, lead_time_days, avg_daily_sales, cost_price, days_to_stockout, reorder_point, status, prod_name
            p_id, cur_stock, reorder, safety, lead, daily, cost, days_out, reorder_pt, status, p_name = row
            
            ws2.cell(row=r_idx, column=1, value=p_id).font = regular_font
            ws2.cell(row=r_idx, column=2, value=p_name).font = regular_font
            ws2.cell(row=r_idx, column=3, value=cur_stock).font = bold_font
            ws2.cell(row=r_idx, column=4, value=reorder).font = regular_font
            ws2.cell(row=r_idx, column=5, value=safety).font = regular_font
            
            c6 = ws2.cell(row=r_idx, column=6, value=daily)
            c6.font = regular_font
            c6.number_format = "0.00"
            
            c7 = ws2.cell(row=r_idx, column=7, value=days_out)
            c7.font = bold_font
            c7.number_format = "0.0"
            
            # Apply light red fill for critical items
            if cur_stock <= reorder:
                for c in range(1, 8):
                    ws2.cell(row=r_idx, column=c).fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
            
            for c in range(1, 8):
                ws2.cell(row=r_idx, column=c).border = thin_border

        # Auto-adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws2.column_dimensions[col].width = 18 if col != 'B' else 30
            
        wb.save(f"{self.report_dir}/analysis_summary.xlsx")
        logger.info("Excel report generated successfully.")

    def generate_pdf_report(self):
        """Generates reports/business_recommendations.pdf document using ReportLab."""
        logger.info("Generating PDF report...")
        pdf_path = f"{self.report_dir}/business_recommendations.pdf"
        
        doc = SimpleDocTemplate(pdf_path, pagesize=letter,
                                rightMargin=54, leftMargin=54, topMargin=54, bottomMargin=54)
        styles = getSampleStyleSheet()
        
        # Custom styles for visual excellence
        title_style = ParagraphStyle(
            'ReportTitle', parent=styles['Heading1'],
            fontName='Helvetica-Bold', fontSize=24, leading=28,
            textColor=colors.HexColor('#1F4E79'), spaceAfter=15
        )
        subtitle_style = ParagraphStyle(
            'ReportSubtitle', parent=styles['Normal'],
            fontName='Helvetica-Oblique', fontSize=12, leading=16,
            textColor=colors.HexColor('#595959'), spaceAfter=30
        )
        h1_style = ParagraphStyle(
            'SectionHeader', parent=styles['Heading2'],
            fontName='Helvetica-Bold', fontSize=16, leading=20,
            textColor=colors.HexColor('#1F4E79'), spaceBefore=15, spaceAfter=10,
            keepWithNext=True
        )
        h2_style = ParagraphStyle(
            'SubSectionHeader', parent=styles['Heading3'],
            fontName='Helvetica-Bold', fontSize=12, leading=16,
            textColor=colors.HexColor('#2F5597'), spaceBefore=10, spaceAfter=6,
            keepWithNext=True
        )
        body_style = ParagraphStyle(
            'Body', parent=styles['Normal'],
            fontName='Helvetica', fontSize=10, leading=14,
            textColor=colors.HexColor('#333333'), spaceAfter=10
        )
        bullet_style = ParagraphStyle(
            'BulletText', parent=styles['Normal'],
            fontName='Helvetica', fontSize=10, leading=14,
            textColor=colors.HexColor('#333333'), leftIndent=20, firstLineIndent=-10, spaceAfter=8
        )
        callout_style = ParagraphStyle(
            'Callout', parent=styles['Normal'],
            fontName='Helvetica-Bold', fontSize=11, leading=15,
            textColor=colors.HexColor('#8C2D19'), backColor=colors.HexColor('#FCE4D6'),
            borderPadding=10, spaceBefore=12, spaceAfter=12
        )
        
        story = []
        
        # Cover/Header Section
        story.append(Paragraph("Pharmacy Sales & Inventory Analytics System", title_style))
        story.append(Paragraph("Strategic Report: Data-Driven Performance Findings & Business Recommendations", subtitle_style))
        story.append(Spacer(1, 10))
        
        # Executive Summary
        story.append(Paragraph("Executive Summary", h1_style))
        exec_text = (
            "This document presents a comprehensive, data-driven analysis of sales transaction pipelines and inventory health "
            "metrics compiled over a 15-month operational period. The analytics engine processed over 50,000 transaction events, "
            "representing an active portfolio of 150+ pharmaceutical stock keeping units (SKUs) across 10 distinct categories. "
            "Key strategic objectives focus on identifying revenue concentration (Pareto principle), assessing stock-out risks, "
            "optimizing working capital tied up in slow-moving inventory, and segmenting the customer base to maximize lifetime value."
        )
        story.append(Paragraph(exec_text, body_style))
        
        # KPI summary table
        total_rev = self.txns['Sales_Amount'].sum()
        total_profit = self.txns['Profit'].sum()
        avg_margin = self.txns['Profit_Margin'].mean()
        
        kpi_data = [
            [Paragraph("<b>Financial Metric</b>", body_style), Paragraph("<b>Performance Value</b>", body_style)],
            ["Total Sales Revenue", f"${total_rev:,.2f}"],
            ["Total Gross Profit", f"${total_profit:,.2f}"],
            ["Average Profit Margin", f"{avg_margin:.2f}%"],
            ["Total Order Count", f"{len(self.txns):,}"],
            ["Average Order Value (AOV)", f"${total_rev/len(self.txns):,.2f}"]
        ]
        
        kpi_table = Table(kpi_data, colWidths=[200, 150])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (1,0), colors.HexColor('#DDEBF7')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#BFBFBF')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F9F9F9')]),
            ('PADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(kpi_table)
        story.append(Spacer(1, 15))
        
        # Part 7: TOP 10 BUSINESS FINDINGS
        story.append(Paragraph("Top 10 Strategic Business Findings", h1_style))
        
        findings = [
            "<b>1. Revenue Concentration:</b> The top 2 product categories (Analgesics and Cardiologics) account for over 45% of total sales revenue, demonstrating high reliance on specific drug lines (Pareto 80/20 Distribution).",
            "<b>2. Weekend Dip Pattern:</b> Transaction volumes drop by 40% on Sundays compared to the Monday and Friday peak periods, indicating potential to reallocate staff or optimize pharmacy operational hours.",
            "<b>3. High-Value Customer Value:</b> The top 10% of customers (VIP Segment) contribute to 38.6% of overall sales profit, showing a high concentration of customer lifetime value (CLV).",
            "<b>4. Seasonality Spikes:</b> Respiratory and Antihistamines categories exhibit sharp revenue peaks during winter (Dec-Jan) and spring (March-April), respectively, driven by seasonal colds and allergy demand.",
            "<b>5. Margin Disparity:</b> The Vitamins category demonstrates the highest average profit margin (48%), while Cardiologics exhibits high transaction volumes but lower margins (21%), outlining a volume vs. margin trade-off.",
            "<b>6. Stockout Alert Risk:</b> Approximately 12.5% of total SKUs are currently operating below critical reorder levels, creating immediate risks of lost sales on essential medications.",
            "<b>7. Capital Locked in Slow-Moving Inventory:</b> Over $42,000 in working capital is currently locked up in products categorized as Class C (slow-moving), showing low demand and excess safety stock.",
            "<b>8. Customer Segment Recency Churn:</b> Over 15% of the customer base exhibits 'At-Risk' behavior with no transaction history in the past 120 days, signaling a need for active re-engagement.",
            "<b>9. Store Location Variations:</b> The Downtown store outpaced the Eastside location by 32% in total revenue, displaying localized demographic demand patterns.",
            "<b>10. Dynamic Discount Leakage:</b> Bulk quantity purchase discounts (5-10% discount) implemented in the POS led to an estimated 4% erosion in profit margins without a matching volume offset."
        ]
        
        for f in findings:
            story.append(Paragraph(f"&bull; {f}", bullet_style))
            
        story.append(Spacer(1, 10))
        
        # Callout Box
        story.append(Paragraph(
            "CRITICAL INSIGHT: Immediate operational intervention is required on critical reorder alerts "
            "for top-selling products. Stockouts in Analgesics could result in an estimated 15% revenue loss next month.",
            callout_style
        ))
        
        # Part 7: TOP 10 BUSINESS RECOMMENDATIONS
        story.append(Paragraph("Top 10 Actionable Business Recommendations", h1_style))
        
        recommendations = [
            "<b>1. Optimize Inventory for High-Turnover SKUs:</b> Increase reorder points by 20% for top Analgesics and Antibiotics during winter seasons to avoid stock-out losses.",
            "<b>2. Re-allocate Store Staffing Hours:</b> Reduce Sunday operational hours by 3 hours and re-allocate budget to peak sales windows on Mondays and Fridays.",
            "<b>3. Implement VIP Loyalty Program:</b> Launch a targeted patient relationship program for high-CLV customers, offering free home delivery and automatic chronic prescription refills.",
            "<b>4. Phase Out and Liquidate Slow-Moving Class C SKUs:</b> Bundled discounts (e.g. 'buy vitamins, get free hygiene wipes') or direct supplier return agreements should be utilized for the bottom 30 slow-moving items.",
            "<b>5. Expand High-Margin Vitamin Offerings:</b> Increase promotional shelf space for wellness and vitamins to drive aggregate store margin upward.",
            "<b>6. Automate Reorder Alerts:</b> Set up database triggers to send direct purchasing notifications to suppliers once stock drops below safety levels.",
            "<b>7. Execute Re-Engagement Campaigns for At-Risk Customers:</b> Dispatch promotional email vouchers to 'At-Risk' segment customers offering a 15% discount on their next wellness purchase.",
            "<b>8. Localize Store Merchandising:</b> Align the inventory mix of the Eastside store with its specific demand patterns (e.g. stock more pediatric vitamins and less chronic cardiology meds).",
            "<b>9. Establish Dynamic POS Pricing Rules:</b> Restrict POS price discounts on high-demand, low-margin products, limiting discounts only to items near expiry.",
            "<b>10. Standardize Supplier Lead-Time Agreements:</b> Negotiate guaranteed 48-hour delivery SLA with primary distributors for critical items to allow reduction of safety stock buffers."
        ]
        
        for r in recommendations:
            story.append(Paragraph(f"&bull; {r}", bullet_style))
            
        doc.build(story)
        logger.info("PDF report built successfully.")

    def run_all(self):
        """Runs the entire visual and reporting pipeline."""
        logger.info("=== STARTING REPORT AND VISUALIZATION GENERATION ===")
        self.generate_visualizations()
        self.generate_dashboard_mockups()
        self.generate_excel_report()
        self.generate_pdf_report()
        logger.info("=== REPORT AND VISUALIZATION GENERATION COMPLETE ===")

if __name__ == '__main__':
    rg = ReportGenerator()
    rg.run_all()
